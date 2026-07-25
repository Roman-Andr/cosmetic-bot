"""Reliable Telegram handlers for products, support, codes and sales."""

from __future__ import annotations

import logging
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from uuid import UUID

from aiogram import Bot, F, Router
from aiogram.filters import Command, CommandObject, CommandStart, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.types import (
    BufferedInputFile,
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    Message,
)
from openpyxl import Workbook
from sqlalchemy import delete, func, select
from sqlalchemy.orm import selectinload

from app.bot.keyboards import (
    customer_menu,
    loyalty_web_app_keyboard,
    owner_menu,
    sale_confirmation_keyboard,
    sale_product_keyboard,
    support_keyboard,
)
from app.bot.states import OwnerStates, SaleStates
from app.core.config import get_settings
from app.db.session import SessionLocal
from app.models import (
    AdminRole,
    AdminUser,
    AuditEvent,
    BlockedUser,
    ContactShare,
    Customer,
    LoyaltyAccount,
    LoyaltyCode,
    LoyaltyTierRule,
    Product,
    Purchase,
    SupportDialog,
    SupportDialogStatus,
    SupportForward,
)
from app.services.loyalty import (
    InvalidCodeError,
    LoyaltyError,
    LoyaltyService,
    ProductSelection,
    money,
)

logger = logging.getLogger(__name__)
router = Router(name="customer-and-admin-workflows")
settings = get_settings()


def normalize_phone(phone: str) -> str:
    """Keep the same normalized phone representation used by API registration."""
    return f"+{''.join(character for character in phone if character.isdigit())}"


def mask_phone(phone: str) -> str:
    """Show sales administrators enough context without exposing all digits."""
    if len(phone) <= 5:
        return phone
    return f"{phone[:4]}{'*' * max(1, len(phone) - 6)}{phone[-2:]}"


async def get_admin_role(telegram_user_id: int) -> AdminRole | None:
    """Load a current active administrator role."""
    async with SessionLocal() as session:
        admin = await session.get(AdminUser, telegram_user_id)
        if admin is None or not admin.is_active:
            return None
        return admin.role


async def is_blocked(telegram_user_id: int) -> bool:
    """Check the persistent owner-managed support blocklist."""
    async with SessionLocal() as session:
        return await session.get(BlockedUser, telegram_user_id) is not None


async def require_sales(message: Message) -> bool:
    """Reject a command unless the sender is an active sales user or owner."""
    if message.from_user is None or await get_admin_role(message.from_user.id) is None:
        await message.answer("Эта команда доступна только администраторам продаж.")
        return False
    return True


async def require_owner_message(message: Message) -> bool:
    """Reject a command unless the sender is the configured owner."""
    if (
        message.from_user is None
        or await get_admin_role(message.from_user.id) is not AdminRole.OWNER
    ):
        await message.answer("Эта команда доступна только главному администратору.")
        return False
    return True


@router.message(CommandStart())
async def start(message: Message, command: CommandObject) -> None:
    """Keep product deep links and add the generic loyalty entry link."""
    if message.from_user is None:
        return
    payload = (command.args or "").strip()
    if payload == "loyalty":
        await message.answer(
            "Добро пожаловать в программу лояльности Velina Cosmetic.",
            reply_markup=loyalty_web_app_keyboard(settings.public_base_url),
        )
        return
    if not payload:
        await message.answer(
            "Откройте программу лояльности или перейдите по ссылке на товар.",
            reply_markup=customer_menu(),
        )
        return

    async with SessionLocal() as session:
        product = await session.scalar(
            select(Product).where(Product.external_id == payload, Product.is_active.is_(True))
        )
    if product is None:
        await message.answer("Товар не найден. Откройте программу лояльности через меню.")
        return

    text = f"{product.title}\n\nСтоимость: {product.current_price or 'уточняйте'} BYN"
    keyboard_rows: list[list[InlineKeyboardButton]] = []
    if product.url:
        keyboard_rows.append([InlineKeyboardButton(text="Добавить в корзину", url=product.url)])
    keyboard_rows.append(
        [InlineKeyboardButton(text="Нужна помощь", callback_data=f"support:{product.id}")]
    )
    keyboard = InlineKeyboardMarkup(inline_keyboard=keyboard_rows)
    if product.photo_url:
        await message.answer_photo(product.photo_url, caption=text, reply_markup=keyboard)
    else:
        await message.answer(text, reply_markup=keyboard)


@router.message(F.contact)
async def contact_shared(message: Message) -> None:
    """Persist a phone number only when Telegram confirms it belongs to the sender."""
    if message.from_user is None or message.contact is None:
        return
    if message.contact.user_id != message.from_user.id:
        await message.answer("Пожалуйста, поделитесь своим собственным номером телефона.")
        return
    phone = normalize_phone(message.contact.phone_number)
    async with SessionLocal() as session:
        share = await session.get(ContactShare, message.from_user.id)
        if share is None:
            session.add(ContactShare(telegram_user_id=message.from_user.id, phone=phone))
        else:
            share.phone = phone
            share.shared_at = datetime.now(UTC)
        await session.commit()
    await message.answer("Номер получен. Вернитесь в Mini App и завершите регистрацию.")


@router.message(F.text == "Программа лояльности")
@router.message(Command("loyalty"))
async def open_loyalty(message: Message) -> None:
    """Open the Mini App from a standard bot menu action."""
    await message.answer(
        "Ваш баланс, история покупок и регистрация доступны в Mini App.",
        reply_markup=loyalty_web_app_keyboard(settings.public_base_url),
    )


@router.message(Command("work"))
async def open_sales_workplace(message: Message) -> None:
    """Give sales users a direct, role-checked entry point to their Mini App workspace."""
    if not await require_sales(message):
        return
    await message.answer(
        "Откройте рабочий кабинет для оформления покупки.",
        reply_markup=loyalty_web_app_keyboard(settings.public_base_url),
    )


@router.message(F.text == "Получить код")
@router.message(Command("code"))
async def get_customer_code(message: Message) -> None:
    """Provide the same one-hour code flow as the Mini App."""
    if message.from_user is None:
        return
    async with SessionLocal() as session:
        account_id = await session.scalar(
            select(LoyaltyAccount.id)
            .join(Customer, Customer.id == LoyaltyAccount.customer_id)
            .where(Customer.telegram_user_id == message.from_user.id)
        )
        if account_id is None:
            await message.answer(
                "Сначала зарегистрируйтесь в программе лояльности.",
                reply_markup=loyalty_web_app_keyboard(settings.public_base_url),
            )
            return
        code, expires_at = await LoyaltyService(settings).generate_code(session, account_id)
    await message.answer(
        f"Ваш код: <code>{code}</code>\nДействует до {expires_at.astimezone().strftime('%H:%M')}.",
        parse_mode="HTML",
    )


@router.callback_query(
    F.data.startswith("support:")
    & ~F.data.startswith("support:end:")
    & ~F.data.startswith("support:block:")
    & ~F.data.startswith("support:unblock:")
)
async def open_support(callback: CallbackQuery) -> None:
    """Open or reuse a persistent support dialog for the selected product."""
    if callback.from_user is None or callback.data is None:
        return
    if await is_blocked(callback.from_user.id):
        await callback.answer("Обращения для этого аккаунта ограничены.", show_alert=True)
        return
    product_token = callback.data.removeprefix("support:")
    async with SessionLocal() as session:
        product_external_id: str | None = None
        if product_token != "general":
            try:
                product_id = UUID(product_token)
            except ValueError:
                await callback.answer(
                    "Карточка товара устарела. Откройте её заново.", show_alert=True
                )
                return
            product = await session.get(Product, product_id)
            if product is None or not product.is_active:
                await callback.answer("Товар больше недоступен.", show_alert=True)
                return
            product_external_id = product.external_id
        dialog = await session.scalar(
            select(SupportDialog).where(
                SupportDialog.customer_telegram_id == callback.from_user.id,
                SupportDialog.status == SupportDialogStatus.OPEN,
            )
        )
        if dialog is None:
            dialog = SupportDialog(
                customer_telegram_id=callback.from_user.id,
                customer_name=callback.from_user.full_name,
                product_external_id=product_external_id,
            )
            session.add(dialog)
            await session.commit()
    await callback.answer("Напишите ваш вопрос.")
    if callback.message:
        await callback.message.answer(
            "Напишите ваш вопрос — главный администратор ответит в этом чате."
        )


@router.message(Command("sale"))
async def start_sale(message: Message, state: FSMContext) -> None:
    """Start a role-protected sale draft in Redis."""
    if not await require_sales(message):
        return
    await state.clear()
    await state.set_state(SaleStates.buyer_code)
    await message.answer("Введите 6-значный код покупателя.")


@router.message(SaleStates.buyer_code, F.text)
async def sale_buyer_code(message: Message, state: FSMContext) -> None:
    """Validate code shape before requesting the order total."""
    code = (message.text or "").strip()
    if not code.isdigit() or len(code) != 6:
        await message.answer("Код должен состоять из шести цифр.")
        return
    await state.update_data(buyer_code=code)
    await state.set_state(SaleStates.total_amount)
    await message.answer("Введите полную сумму покупки в BYN.")


@router.message(SaleStates.total_amount, F.text)
async def sale_total_amount(message: Message, state: FSMContext) -> None:
    """Validate the full amount before optional catalogue selection."""
    try:
        total = money((message.text or "").strip().replace(",", "."))
    except (InvalidOperation, ValueError):
        await message.answer("Введите корректную сумму, например 49.90.")
        return
    if total <= 0:
        await message.answer("Сумма должна быть больше нуля.")
        return
    await state.update_data(total_amount=str(total), product_ids=[])
    await state.set_state(SaleStates.products)
    await message.answer(
        "Найдите товар текстом или завершите оформление без товаров.",
        reply_markup=sale_product_keyboard([]),
    )


@router.message(SaleStates.products, F.text)
async def sale_product_search(message: Message, state: FSMContext) -> None:
    """Search locally synchronized products and add them through compact callbacks."""
    query = (message.text or "").strip()
    async with SessionLocal() as session:
        products = list(
            (
                await session.scalars(
                    select(Product)
                    .where(
                        Product.is_active.is_(True),
                        Product.title.ilike(f"%{query}%"),
                    )
                    .order_by(Product.title)
                    .limit(10)
                )
            ).all()
        )
    if not products:
        await message.answer("Товары не найдены. Попробуйте другой запрос или завершите выбор.")
        return
    keyboard = InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text=product.title[:60], callback_data=f"sale:add:{product.id}")]
            for product in products
        ]
    )
    await message.answer("Выберите товар:", reply_markup=keyboard)


@router.callback_query(SaleStates.products, F.data.startswith("sale:add:"))
async def sale_add_product(callback: CallbackQuery, state: FSMContext) -> None:
    """Add a selected optional product and keep the administrator in search mode."""
    if callback.data is None:
        return
    product_id = callback.data.removeprefix("sale:add:")
    try:
        UUID(product_id)
    except ValueError:
        await callback.answer("Некорректный товар.", show_alert=True)
        return
    data = await state.get_data()
    product_ids = list(dict.fromkeys([*data.get("product_ids", []), product_id]))
    await state.update_data(product_ids=product_ids)
    await callback.answer("Товар добавлен")
    if callback.message:
        await callback.message.answer(
            f"Выбрано товаров: {len(product_ids)}. Ищите следующий или завершите выбор.",
            reply_markup=sale_product_keyboard(product_ids),
        )


@router.callback_query(SaleStates.products, F.data.in_({"sale:skip", "sale:finish"}))
async def sale_preview(callback: CallbackQuery, state: FSMContext) -> None:
    """Show all irreversible sale details before the final confirmation button."""
    data = await state.get_data()
    product_ids = [] if callback.data == "sale:skip" else data.get("product_ids", [])
    async with SessionLocal() as session:
        try:
            preview = await LoyaltyService(settings).preview_purchase(
                session,
                buyer_code=data["buyer_code"],
                total_amount=data["total_amount"],
            )
            product_titles = list(
                (
                    await session.scalars(
                        select(Product.title).where(
                            Product.id.in_([UUID(value) for value in product_ids])
                        )
                    )
                ).all()
            )
        except (InvalidCodeError, LoyaltyError) as exc:
            await state.clear()
            await callback.answer("Код больше не действует.", show_alert=True)
            if callback.message:
                await callback.message.answer(f"Продажа не оформлена: {exc}")
            return
    await state.update_data(product_ids=product_ids)
    await state.set_state(SaleStates.confirmation)
    product_text = ", ".join(product_titles) if product_titles else "не выбраны"
    cashback_label = (
        "кешбэк ко дню рождения"
        if preview.cashback_source.value == "birthday"
        else "уровень лояльности"
    )
    text = (
        "Проверьте продажу:\n\n"
        f"Клиент: {preview.customer_name}, {mask_phone(preview.customer_phone)}\n"
        f"Сумма: {preview.total_amount} BYN\n"
        f"Списать бонусов: {preview.redeemed} BYN\n"
        f"К оплате деньгами: {preview.cash_paid} BYN\n"
        f"Начислить: {preview.accrued} бонусов ({preview.cashback_percent}%, {cashback_label})\n"
        f"Товары: {product_text}"
    )
    await callback.answer()
    if callback.message:
        await callback.message.answer(text, reply_markup=sale_confirmation_keyboard())


@router.callback_query(SaleStates.confirmation, F.data == "sale:confirm")
async def sale_confirm(callback: CallbackQuery, state: FSMContext) -> None:
    """Atomically create the purchase, ledger rows, and notification outbox record."""
    if callback.from_user is None:
        return
    data = await state.get_data()
    async with SessionLocal() as session:
        product_ids = [UUID(value) for value in data.get("product_ids", [])]
        external_ids = list(
            (
                await session.scalars(
                    select(Product.external_id).where(Product.id.in_(product_ids))
                )
            ).all()
        )
        try:
            result = await LoyaltyService(settings).record_purchase(
                session,
                buyer_code=data["buyer_code"],
                recorded_by_telegram_id=callback.from_user.id,
                total_amount=data["total_amount"],
                selected_products=[ProductSelection(external_id=value) for value in external_ids],
            )
        except (InvalidCodeError, LoyaltyError) as exc:
            await state.clear()
            await callback.answer("Продажа не оформлена.", show_alert=True)
            if callback.message:
                await callback.message.answer(str(exc))
            return
    await state.clear()
    await callback.answer("Продажа зафиксирована")
    if callback.message:
        cashback_label = ""
        if result.cashback_source.value == "birthday":
            cashback_label = " по акции ко дню рождения"
        await callback.message.answer(
            "Продажа зафиксирована. "
            f"Списано: {result.redeemed}; начислено: {result.accrued}; "
            f"баланс клиента: {result.balance_after}{cashback_label}."
        )


@router.callback_query(
    StateFilter(SaleStates.products, SaleStates.confirmation), F.data == "sale:cancel"
)
async def sale_cancel(callback: CallbackQuery, state: FSMContext) -> None:
    """Drop a draft safely before it affects balances or turnover."""
    await state.clear()
    await callback.answer("Черновик отменён")
    if callback.message:
        await callback.message.answer("Оформление продажи отменено.")


@router.message(Command("admin"))
async def owner_panel(message: Message) -> None:
    """Show owner controls without exposing them to sales-only administrators."""
    if not await require_owner_message(message):
        return
    await message.answer(
        "Панель главного администратора\n\n"
        "Команды: /stats, /find, /admins, /addsales, /tiers, "
        "/exportcustomers, /exportpurchases",
        reply_markup=owner_menu(settings.public_base_url),
    )


async def owner_stats_text() -> str:
    """Build one all-time snapshot for both the bot button and the /stats command."""
    async with SessionLocal() as session:
        registrations = await session.scalar(select(func.count(Customer.id)))
        purchase_count = await session.scalar(select(func.count(Purchase.id)))
        turnover = await session.scalar(
            select(func.coalesce(func.sum(LoyaltyAccount.lifetime_turnover), 0))
        )
        liability = await session.scalar(
            select(func.coalesce(func.sum(LoyaltyAccount.current_balance), 0))
        )
    return (
        "Статистика за всё время:\n"
        f"Клиентов: {registrations or 0}\n"
        f"Покупок: {purchase_count or 0}\n"
        f"Оборот: {turnover or 0} BYN\n"
        f"Бонусные обязательства: {liability or 0} BYN"
    )


@router.callback_query(F.data == "owner:stats")
async def owner_stats(callback: CallbackQuery) -> None:
    """Show essential all-time metrics directly in the bot."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await callback.answer()
    if callback.message:
        await callback.message.answer(await owner_stats_text())


@router.message(Command("stats"))
async def owner_stats_command(message: Message) -> None:
    """Expose the owner dashboard summary directly as a discoverable command."""
    if not await require_owner_message(message):
        return
    await message.answer(await owner_stats_text())


@router.message(Command("admins"))
async def owner_administrators_command(message: Message) -> None:
    """List the current owner and sales accounts without requiring the Mini App."""
    if not await require_owner_message(message):
        return
    async with SessionLocal() as session:
        administrators = list(
            (await session.scalars(select(AdminUser).order_by(AdminUser.created_at))).all()
        )
    if not administrators:
        await message.answer("Администраторы пока не добавлены.")
        return
    lines = [
        f"• <code>{admin.telegram_user_id}</code> — "
        f"{'главный' if admin.role is AdminRole.OWNER else 'продажи'}"
        for admin in administrators
        if admin.is_active
    ]
    await message.answer("Администраторы:\n" + "\n".join(lines), parse_mode="HTML")


@router.message(Command("find"))
async def owner_find_customer_command(message: Message, command: CommandObject) -> None:
    """Find loyalty customers by name, phone, or an active six-digit code in the bot."""
    if not await require_owner_message(message):
        return
    query = (command.args or "").strip()
    if len(query) < 2:
        await message.answer(
            "Используйте: <code>/find ФИО, телефон или код</code>",
            parse_mode="HTML",
        )
        return
    condition = Customer.full_name.ilike(f"%{query}%") | Customer.phone.ilike(f"%{query}%")
    async with SessionLocal() as session:
        rows = list(
            (
                await session.execute(
                    select(Customer, LoyaltyAccount)
                    .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                    .where(condition)
                    .order_by(Customer.full_name)
                    .limit(20)
                )
            ).all()
        )
        if query.isdigit() and len(query) == 6:
            from app.services.loyalty import code_digest

            code = await session.scalar(
                select(LoyaltyCode)
                .where(LoyaltyCode.code_digest == code_digest(query, settings.loyalty_code_pepper))
                .limit(1)
            )
            if code is not None:
                row = await session.execute(
                    select(Customer, LoyaltyAccount)
                    .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                    .where(LoyaltyAccount.id == code.account_id)
                )
                found = row.one_or_none()
                rows = [found] if found is not None else rows
    if not rows:
        await message.answer("Клиенты не найдены.")
        return
    lines = [
        f"• {customer.full_name} — {customer.phone}\n"
        f"  Баланс: {account.current_balance} BYN · Оборот: {account.lifetime_turnover} BYN"
        for customer, account in rows
    ]
    await message.answer("Результаты поиска:\n" + "\n".join(lines))


@router.message(Command("exportcustomers"))
async def owner_export_customers_command(message: Message) -> None:
    """Send the owner the same customer XLSX export available in the Mini App."""
    if not await require_owner_message(message):
        return
    async with SessionLocal() as session:
        rows = list(
            (
                await session.execute(
                    select(Customer, LoyaltyAccount)
                    .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                    .order_by(Customer.created_at.desc())
                )
            ).all()
        )
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Клиенты")
    sheet.append(["ФИО", "Телефон", "Дата регистрации", "Баланс", "Оборот"])
    for customer, account in rows:
        sheet.append(
            [
                customer.full_name,
                customer.phone,
                customer.created_at.isoformat(),
                account.current_balance,
                account.lifetime_turnover,
            ]
        )
    output = BytesIO()
    workbook.save(output)
    await message.answer_document(
        BufferedInputFile(output.getvalue(), filename="customers.xlsx"),
        caption="Выгрузка клиентов",
    )


@router.message(Command("exportpurchases"))
async def owner_export_purchases_command(message: Message) -> None:
    """Send the owner the same immutable purchase report available in the Mini App."""
    if not await require_owner_message(message):
        return
    async with SessionLocal() as session:
        rows = list(
            (
                await session.execute(
                    select(Purchase, Customer)
                    .join(Customer, Customer.id == Purchase.customer_id)
                    .options(selectinload(Purchase.items))
                    .order_by(Purchase.created_at.desc())
                )
            )
            .unique()
            .all()
        )
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Покупки")
    sheet.append(["Дата", "Клиент", "Сумма", "Списано", "Начислено", "Кешбэк", "Товары"])
    for purchase, customer in rows:
        source = "День рождения" if purchase.cashback_source.value == "birthday" else "Уровень"
        sheet.append(
            [
                purchase.created_at.isoformat(),
                customer.full_name,
                purchase.total_amount,
                purchase.bonus_redeemed,
                purchase.cashback_accrued,
                f"{source} · {purchase.cashback_percent}%",
                ", ".join(item.title_snapshot for item in purchase.items),
            ]
        )
    output = BytesIO()
    workbook.save(output)
    await message.answer_document(
        BufferedInputFile(output.getvalue(), filename="purchases.xlsx"),
        caption="Выгрузка покупок",
    )


@router.callback_query(F.data == "owner:add-sales")
async def owner_add_sales_start(callback: CallbackQuery, state: FSMContext) -> None:
    """Request a sales administrator's numeric Telegram ID."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    await state.set_state(OwnerStates.sales_admin_id)
    await callback.answer()
    if callback.message:
        await callback.message.answer("Введите Telegram ID администратора продаж.")


@router.message(Command("addsales"))
async def owner_add_sales_command(message: Message, state: FSMContext) -> None:
    """Expose the same sales-admin workflow through an owner command."""
    if not await require_owner_message(message):
        return
    await state.set_state(OwnerStates.sales_admin_id)
    await message.answer("Введите Telegram ID администратора продаж.")


@router.message(OwnerStates.sales_admin_id, F.text)
async def owner_add_sales_finish(message: Message, state: FSMContext) -> None:
    """Grant sales-only access from the owner bot panel."""
    if not await require_owner_message(message):
        return
    if message.from_user is None:
        return
    try:
        telegram_user_id = int((message.text or "").strip())
        if telegram_user_id <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите положительный числовой Telegram ID.")
        return
    async with SessionLocal() as session:
        existing = await session.get(AdminUser, telegram_user_id)
        if existing is None:
            session.add(
                AdminUser(
                    telegram_user_id=telegram_user_id,
                    role=AdminRole.SALES,
                    added_by_telegram_id=message.from_user.id,
                )
            )
        elif existing.role is not AdminRole.OWNER:
            existing.role = AdminRole.SALES
            existing.is_active = True
            existing.added_by_telegram_id = message.from_user.id
        else:
            await message.answer("Этот ID уже принадлежит главному администратору.")
            return
        await session.commit()
    await state.clear()
    await message.answer("Администратор продаж добавлен.")


@router.callback_query(F.data == "owner:tiers")
async def owner_tiers_start(callback: CallbackQuery, state: FSMContext) -> None:
    """Let the owner update the cashback table in the bot as well as the Mini App."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    async with SessionLocal() as session:
        tiers = list(
            (
                await session.scalars(
                    select(LoyaltyTierRule)
                    .where(LoyaltyTierRule.is_active.is_(True))
                    .order_by(LoyaltyTierRule.minimum_turnover)
                )
            ).all()
        )
    current = ", ".join(f"{tier.minimum_turnover}:{tier.cashback_percent}" for tier in tiers)
    await state.set_state(OwnerStates.tier_rules)
    await callback.answer()
    if callback.message:
        await callback.message.answer(
            "Отправьте уровни в формате <code>порог:процент, порог:процент</code>.\n"
            "Первый порог должен быть 0. Изменение повлияет только на будущие начисления.\n"
            f"Сейчас: <code>{current}</code>",
            parse_mode="HTML",
        )


@router.message(Command("tiers"))
async def owner_tiers_command(message: Message, state: FSMContext) -> None:
    """Expose the tier workflow through a discoverable owner command."""
    if not await require_owner_message(message):
        return
    await state.set_state(OwnerStates.tier_rules)
    await message.answer(
        "Отправьте уровни, например: <code>0:3, 1000:5, 2000:7</code>",
        parse_mode="HTML",
    )


def parse_tier_rules(raw: str) -> list[tuple[Decimal, Decimal]]:
    """Parse a compact owner-only tier table and reject ambiguous configurations."""
    chunks = [chunk.strip() for chunk in raw.split(",") if chunk.strip()]
    if not 1 <= len(chunks) <= 10:
        raise ValueError("Количество уровней должно быть от 1 до 10")
    rules: list[tuple[Decimal, Decimal]] = []
    for chunk in chunks:
        threshold_text, separator, percent_text = chunk.partition(":")
        if not separator:
            raise ValueError("Используйте формат порог:процент")
        threshold = money(threshold_text.strip().replace(",", "."))
        percent = Decimal(percent_text.strip().replace(",", ".")).quantize(Decimal("0.01"))
        if threshold < 0 or not Decimal("0") <= percent <= Decimal("100"):
            raise ValueError("Порог должен быть неотрицательным, процент — от 0 до 100")
        rules.append((threshold, percent))
    thresholds = [threshold for threshold, _ in rules]
    if (
        thresholds[0] != 0
        or thresholds != sorted(thresholds)
        or len(set(thresholds)) != len(thresholds)
    ):
        raise ValueError("Первый порог должен быть 0, остальные — строго возрастать")
    return rules


@router.message(OwnerStates.tier_rules, F.text)
async def owner_tiers_finish(message: Message, state: FSMContext) -> None:
    """Atomically replace tier rules and retain an owner audit trail."""
    if not await require_owner_message(message) or message.from_user is None:
        return
    try:
        rules = parse_tier_rules(message.text or "")
    except (InvalidOperation, ValueError) as exc:
        await message.answer(f"Не удалось сохранить уровни: {exc}")
        return
    async with SessionLocal() as session:
        previous = list(
            (
                await session.scalars(
                    select(LoyaltyTierRule).where(LoyaltyTierRule.is_active.is_(True))
                )
            ).all()
        )
        await session.execute(delete(LoyaltyTierRule))
        session.add_all(
            LoyaltyTierRule(
                minimum_turnover=threshold,
                cashback_percent=percent,
                updated_by_telegram_id=message.from_user.id,
            )
            for threshold, percent in rules
        )
        session.add(
            AuditEvent(
                actor_telegram_id=message.from_user.id,
                event_type="loyalty_tiers_replaced",
                target_type="loyalty_tier_rules",
                payload={
                    "old": [
                        {
                            "minimum_turnover": str(tier.minimum_turnover),
                            "cashback_percent": str(tier.cashback_percent),
                        }
                        for tier in previous
                    ],
                    "new": [
                        {"minimum_turnover": str(threshold), "cashback_percent": str(percent)}
                        for threshold, percent in rules
                    ],
                },
            )
        )
        await session.commit()
    await state.clear()
    await message.answer("Уровни кешбэка сохранены. Они будут применяться к будущим покупкам.")


@router.callback_query(F.data.startswith("support:end:"))
async def close_support(callback: CallbackQuery) -> None:
    """Close a dialog by opaque UUID rather than parsing human-readable text."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    try:
        dialog_id = UUID((callback.data or "").removeprefix("support:end:"))
    except ValueError:
        await callback.answer("Некорректный диалог", show_alert=True)
        return
    async with SessionLocal() as session:
        dialog = await session.get(SupportDialog, dialog_id)
        if dialog is None:
            await callback.answer("Диалог не найден", show_alert=True)
            return
        dialog.status = SupportDialogStatus.CLOSED
        dialog.closed_at = datetime.now(UTC)
        await session.commit()
    await callback.answer("Диалог завершён")
    if callback.bot is not None:
        await callback.bot.send_message(
            dialog.customer_telegram_id,
            "Диалог завершён. Если потребуется помощь, нажмите «Нужна помощь» на карточке товара.",
        )


@router.callback_query(F.data.startswith("support:block:"))
async def block_support_user(callback: CallbackQuery) -> None:
    """Persistently block a user from opening or continuing support dialogs."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    try:
        dialog_id, customer_id = (callback.data or "").removeprefix("support:block:").split(":")
        UUID(dialog_id)
        customer_telegram_id = int(customer_id)
    except ValueError:
        await callback.answer("Некорректный пользователь", show_alert=True)
        return
    async with SessionLocal() as session:
        if await session.get(BlockedUser, customer_telegram_id) is None:
            session.add(
                BlockedUser(
                    telegram_user_id=customer_telegram_id,
                    blocked_by_telegram_id=callback.from_user.id,
                )
            )
        await session.commit()
    await callback.answer("Пользователь заблокирован")
    if isinstance(callback.message, Message):
        await callback.message.edit_reply_markup(
            reply_markup=support_keyboard(dialog_id, customer_telegram_id, is_blocked=True)
        )


@router.callback_query(F.data.startswith("support:unblock:"))
async def unblock_support_user(callback: CallbackQuery) -> None:
    """Restore support access with the same owner-only control path."""
    if (
        callback.from_user is None
        or await get_admin_role(callback.from_user.id) is not AdminRole.OWNER
    ):
        await callback.answer("Недостаточно прав", show_alert=True)
        return
    try:
        dialog_id, customer_id = (callback.data or "").removeprefix("support:unblock:").split(":")
        UUID(dialog_id)
        customer_telegram_id = int(customer_id)
    except ValueError:
        await callback.answer("Некорректный пользователь", show_alert=True)
        return
    async with SessionLocal() as session:
        blocked = await session.get(BlockedUser, customer_telegram_id)
        if blocked is not None:
            await session.delete(blocked)
            await session.commit()
    await callback.answer("Пользователь разблокирован")
    if isinstance(callback.message, Message):
        await callback.message.edit_reply_markup(
            reply_markup=support_keyboard(dialog_id, customer_telegram_id)
        )


@router.message(F.reply_to_message)
async def owner_reply_to_support(message: Message, bot: Bot) -> None:
    """Route an owner reply by persistent forwarded-message mapping."""
    if message.from_user is None or message.reply_to_message is None:
        return
    if await get_admin_role(message.from_user.id) is not AdminRole.OWNER:
        return
    async with SessionLocal() as session:
        forward = await session.scalar(
            select(SupportForward).where(
                SupportForward.owner_message_id == message.reply_to_message.message_id
            )
        )
        if forward is None:
            return
        dialog = await session.get(SupportDialog, forward.dialog_id)
        if dialog is None or dialog.status is not SupportDialogStatus.OPEN:
            await message.answer("Этот диалог уже закрыт.")
            return
        customer_telegram_id = dialog.customer_telegram_id
    await bot.copy_message(
        chat_id=customer_telegram_id,
        from_chat_id=message.chat.id,
        message_id=message.message_id,
    )


@router.message()
async def forward_customer_support(message: Message, bot: Bot) -> None:
    """Forward every supported customer message in an open dialog to the owner."""
    if message.from_user is None or message.from_user.id == settings.owner_telegram_id:
        return
    if message.text and message.text.startswith("/"):
        return
    if await is_blocked(message.from_user.id):
        return
    async with SessionLocal() as session:
        dialog = await session.scalar(
            select(SupportDialog).where(
                SupportDialog.customer_telegram_id == message.from_user.id,
                SupportDialog.status == SupportDialogStatus.OPEN,
            )
        )
        if dialog is None:
            return
        copied = await bot.copy_message(
            chat_id=settings.owner_telegram_id,
            from_chat_id=message.chat.id,
            message_id=message.message_id,
            reply_markup=support_keyboard(str(dialog.id), message.from_user.id),
        )
        session.add(
            SupportForward(
                dialog_id=dialog.id,
                owner_message_id=copied.message_id,
                customer_message_id=message.message_id,
            )
        )
        await session.commit()
