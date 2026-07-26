"""Owner dashboard, exports, search, configuration and catalogue endpoints."""

from __future__ import annotations

from datetime import UTC, date, datetime, time
from decimal import Decimal
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, HTTPException, Query, Response, status
from openpyxl import Workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import selectinload

from app.api.dependencies import (
    OwnerDependency,
    SalesAdminDependency,
    SessionDependency,
    SettingsDependency,
)
from app.models import (
    AdminRole,
    AdminUser,
    AuditEvent,
    CashbackSource,
    Customer,
    LoyaltyAccount,
    LoyaltyCode,
    LoyaltyTierRule,
    Product,
    Purchase,
)
from app.schemas.admin import (
    AddSalesAdminRequest,
    AdminAccessResponse,
    AdminStatsResponse,
    AdminUserResponse,
    CustomerDetailResponse,
    CustomerSearchResponse,
    ProductResponse,
    PurchasePreviewRequest,
    PurchasePreviewResponse,
    PurchaseRecordRequest,
    PurchaseRecordResponse,
    TierRuleResponse,
    TierRulesUpdateRequest,
)
from app.schemas.loyalty import PurchasePageResponse, PurchaseSummaryResponse
from app.services.loyalty import (
    InvalidCodeError,
    LoyaltyError,
    LoyaltyService,
    ProductNotFoundError,
    ProductSelection,
    code_digest,
)

router = APIRouter(prefix="/admin", tags=["admin"])


def as_utc_boundary(value: date | None, *, end: bool) -> datetime | None:
    """Convert an optional report-date boundary to an inclusive UTC timestamp."""
    if value is None:
        return None
    return datetime.combine(value, time.max if end else time.min, tzinfo=UTC)


def money_or_zero(value: Decimal | None) -> Decimal:
    """Normalise aggregate NULL values for API and XLSX output."""
    return value or Decimal("0.00")


def mask_phone(phone: str) -> str:
    """Show sales staff only enough digits to identify a customer at the till."""
    if len(phone) <= 5:
        return phone
    return f"{phone[:4]}{'*' * max(1, len(phone) - 6)}{phone[-2:]}"


@router.get("/access", response_model=AdminAccessResponse)
async def get_admin_access(admin: SalesAdminDependency) -> AdminAccessResponse:
    """Allow an administrator without a loyalty profile to open the work Mini App."""
    return AdminAccessResponse(role=admin.role)


@router.get("/products", response_model=list[ProductResponse])
async def search_products(
    session: SessionDependency,
    _: SalesAdminDependency,
    query: str = Query(default="", max_length=100),
) -> list[ProductResponse]:
    """List active products and optionally filter them for the sale selector."""
    term = query.strip()
    statement = select(Product).where(Product.is_active.is_(True))
    if term:
        statement = statement.where(
            or_(
                Product.title.ilike(f"%{term}%"),
                Product.external_id.ilike(f"%{term}%"),
            )
        )
    products = list(
        (
            await session.scalars(
                statement.order_by(Product.title).limit(30)
            )
        ).all()
    )
    return [
        ProductResponse(
            external_id=product.external_id,
            title=product.title,
            current_price=product.current_price,
        )
        for product in products
    ]


@router.post("/purchases/preview", response_model=PurchasePreviewResponse)
async def preview_purchase(
    payload: PurchasePreviewRequest,
    session: SessionDependency,
    settings: SettingsDependency,
    _: SalesAdminDependency,
) -> PurchasePreviewResponse:
    """Use the same transactional calculation as the bot before confirming a sale."""
    try:
        preview = await LoyaltyService(settings).preview_purchase(
            session,
            buyer_code=payload.buyer_code,
            total_amount=payload.total_amount,
        )
    except InvalidCodeError as exc:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
    except LoyaltyError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc
    return PurchasePreviewResponse(
        customer_name=preview.customer_name,
        customer_phone_masked=mask_phone(preview.customer_phone),
        current_balance=preview.current_balance,
        total_amount=preview.total_amount,
        bonus_redeemed=preview.redeemed,
        cash_paid=preview.cash_paid,
        cashback_accrued=preview.accrued,
        cashback_percent=preview.cashback_percent,
        cashback_source=preview.cashback_source,
    )


@router.post(
    "/purchases",
    response_model=PurchaseRecordResponse,
    status_code=status.HTTP_201_CREATED,
)
async def record_purchase(
    payload: PurchaseRecordRequest,
    session: SessionDependency,
    settings: SettingsDependency,
    admin: SalesAdminDependency,
) -> PurchaseRecordResponse:
    """Confirm a Mini App sale with the same service used by the bot FSM."""
    try:
        result = await LoyaltyService(settings).record_purchase(
            session,
            buyer_code=payload.buyer_code,
            recorded_by_telegram_id=admin.telegram_user_id,
            total_amount=payload.total_amount,
            selected_products=[
                ProductSelection(external_id=external_id)
                for external_id in dict.fromkeys(payload.product_external_ids)
            ],
        )
    except InvalidCodeError as exc:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=str(exc)) from exc
    except ProductNotFoundError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc
    except LoyaltyError as exc:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=str(exc),
        ) from exc
    return PurchaseRecordResponse(
        purchase_id=result.purchase.id,
        bonus_redeemed=result.redeemed,
        cash_paid=result.purchase.cash_paid,
        cashback_accrued=result.accrued,
        cashback_percent=result.cashback_percent,
        cashback_source=result.cashback_source,
        balance_after=result.balance_after,
    )


@router.get("/customers/search", response_model=list[CustomerSearchResponse])
async def search_customers(
    session: SessionDependency,
    settings: SettingsDependency,
    _: OwnerDependency,
    query: str = Query(min_length=2, max_length=255),
) -> list[CustomerSearchResponse]:
    """Find customers by name, phone, or an active/current six-digit code."""
    condition = or_(
        Customer.full_name.ilike(f"%{query.strip()}%"), Customer.phone.ilike(f"%{query.strip()}%")
    )
    statement = (
        select(Customer, LoyaltyAccount)
        .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
        .where(condition)
        .order_by(Customer.full_name)
        .limit(30)
    )
    rows = list((await session.execute(statement)).all())

    if query.isdigit() and len(query) == 6:
        digest = code_digest(query, settings.loyalty_code_pepper)
        code_rows = list(
            (
                await session.execute(
                    select(Customer, LoyaltyAccount)
                    .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                    .join(LoyaltyCode, LoyaltyCode.account_id == LoyaltyAccount.id)
                    .where(LoyaltyCode.code_digest == digest)
                )
            ).all()
        )
        rows = code_rows or rows

    return [
        CustomerSearchResponse(
            customer_id=customer.id,
            full_name=customer.full_name,
            phone=customer.phone,
            telegram_user_id=customer.telegram_user_id,
            registered_at=customer.created_at,
            current_balance=account.current_balance,
            lifetime_turnover=account.lifetime_turnover,
        )
        for customer, account in rows
    ]


@router.get("/customers/{customer_id}", response_model=CustomerDetailResponse)
async def get_customer(
    customer_id: UUID,
    session: SessionDependency,
    _: OwnerDependency,
) -> CustomerDetailResponse:
    """Return an owner-visible customer card; purchase history has a dedicated endpoint."""
    row = await session.execute(
        select(Customer, LoyaltyAccount)
        .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
        .where(Customer.id == customer_id)
    )
    result = row.one_or_none()
    if result is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    customer, account = result
    return CustomerDetailResponse(
        customer_id=customer.id,
        full_name=customer.full_name,
        phone=customer.phone,
        telegram_user_id=customer.telegram_user_id,
        registered_at=customer.created_at,
        current_balance=account.current_balance,
        lifetime_turnover=account.lifetime_turnover,
        birth_date=customer.birth_date,
        gender=customer.gender.value,
    )


@router.get("/customers/{customer_id}/purchases", response_model=PurchasePageResponse)
async def get_customer_purchases(
    customer_id: UUID,
    session: SessionDependency,
    _: OwnerDependency,
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=20, ge=1, le=50),
) -> PurchasePageResponse:
    """Show the owner a customer's date-and-amount purchase history."""
    customer_exists = await session.scalar(select(Customer.id).where(Customer.id == customer_id))
    if customer_exists is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Customer not found")
    purchases = list(
        (
            await session.scalars(
                select(Purchase)
                .where(Purchase.customer_id == customer_id)
                .order_by(Purchase.created_at.desc())
                .offset(offset)
                .limit(limit + 1)
            )
        ).all()
    )
    return PurchasePageResponse(
        items=[PurchaseSummaryResponse.model_validate(item) for item in purchases[:limit]],
        next_offset=offset + limit if len(purchases) > limit else None,
    )


@router.get("/stats", response_model=AdminStatsResponse)
async def get_stats(
    session: SessionDependency,
    _: OwnerDependency,
    date_from: date | None = None,
    date_to: date | None = None,
) -> AdminStatsResponse:
    """Return the owner dashboard aggregates for an optional period."""
    start = as_utc_boundary(date_from, end=False)
    end = as_utc_boundary(date_to, end=True)
    purchase_filters = []
    customer_filters = []
    if start is not None:
        purchase_filters.append(Purchase.created_at >= start)
        customer_filters.append(Customer.created_at >= start)
    if end is not None:
        purchase_filters.append(Purchase.created_at <= end)
        customer_filters.append(Customer.created_at <= end)

    registrations = await session.scalar(select(func.count(Customer.id)).where(*customer_filters))
    purchase_values = await session.execute(
        select(
            func.count(Purchase.id),
            func.coalesce(func.sum(Purchase.total_amount), 0),
            func.coalesce(func.sum(Purchase.cashback_accrued), 0),
            func.coalesce(func.sum(Purchase.bonus_redeemed), 0),
        ).where(*purchase_filters)
    )
    purchase_count, turnover, accrued, redeemed = purchase_values.one()
    liability = await session.scalar(
        select(func.coalesce(func.sum(LoyaltyAccount.current_balance), 0))
    )

    tiers = list(
        (
            await session.scalars(
                select(LoyaltyTierRule)
                .where(LoyaltyTierRule.is_active.is_(True))
                .order_by(LoyaltyTierRule.minimum_turnover)
            )
        ).all()
    )
    accounts = list((await session.scalars(select(LoyaltyAccount.lifetime_turnover))).all())
    tier_distribution = {f"{tier.cashback_percent}%": 0 for tier in tiers}
    for turnover_value in accounts:
        eligible = [tier for tier in tiers if tier.minimum_turnover <= turnover_value]
        if eligible:
            tier_distribution[f"{eligible[-1].cashback_percent}%"] += 1

    return AdminStatsResponse(
        registrations=registrations or 0,
        purchase_count=purchase_count or 0,
        turnover=money_or_zero(turnover),
        accrued_bonuses=money_or_zero(accrued),
        redeemed_bonuses=money_or_zero(redeemed),
        bonus_liability=money_or_zero(liability),
        tier_distribution=tier_distribution,
    )


@router.get("/tiers", response_model=list[TierRuleResponse])
async def get_tiers(session: SessionDependency, _: OwnerDependency) -> list[TierRuleResponse]:
    """List current owner-configurable loyalty thresholds."""
    tiers = list(
        (
            await session.scalars(
                select(LoyaltyTierRule)
                .where(LoyaltyTierRule.is_active.is_(True))
                .order_by(LoyaltyTierRule.minimum_turnover)
            )
        ).all()
    )
    return [
        TierRuleResponse(
            id=tier.id,
            minimum_turnover=tier.minimum_turnover,
            cashback_percent=tier.cashback_percent,
        )
        for tier in tiers
    ]


@router.put("/tiers", response_model=list[TierRuleResponse])
async def replace_tiers(
    payload: TierRulesUpdateRequest,
    session: SessionDependency,
    owner: OwnerDependency,
) -> list[TierRuleResponse]:
    """Replace active rules and persist the old/new configuration in the audit log."""
    old_rules = list(
        (
            await session.scalars(
                select(LoyaltyTierRule).where(LoyaltyTierRule.is_active.is_(True))
            )
        ).all()
    )
    old_payload = [
        {
            "minimum_turnover": str(rule.minimum_turnover),
            "cashback_percent": str(rule.cashback_percent),
        }
        for rule in old_rules
    ]
    new_payload = [rule.model_dump(mode="json") for rule in payload.rules]
    await session.execute(delete(LoyaltyTierRule))
    new_rules = [
        LoyaltyTierRule(
            minimum_turnover=rule.minimum_turnover,
            cashback_percent=rule.cashback_percent,
            updated_by_telegram_id=owner.telegram_user_id,
        )
        for rule in payload.rules
    ]
    session.add_all(new_rules)
    session.add(
        AuditEvent(
            actor_telegram_id=owner.telegram_user_id,
            event_type="loyalty_tiers_replaced",
            target_type="loyalty_tier_rules",
            payload={"old": old_payload, "new": new_payload},
        )
    )
    await session.commit()
    return [
        TierRuleResponse(
            id=tier.id,
            minimum_turnover=tier.minimum_turnover,
            cashback_percent=tier.cashback_percent,
        )
        for tier in new_rules
    ]


@router.get("/administrators", response_model=list[AdminUserResponse])
async def get_administrators(
    session: SessionDependency,
    _: OwnerDependency,
) -> list[AdminUserResponse]:
    """List owner and sales accounts configured for this bot."""
    administrators = list(
        (await session.scalars(select(AdminUser).order_by(AdminUser.created_at))).all()
    )
    return [
        AdminUserResponse(
            telegram_user_id=admin.telegram_user_id,
            role=admin.role.value,
            is_active=admin.is_active,
            created_at=admin.created_at,
        )
        for admin in administrators
    ]


@router.post(
    "/administrators", response_model=AdminUserResponse, status_code=status.HTTP_201_CREATED
)
async def add_sales_administrator(
    payload: AddSalesAdminRequest,
    session: SessionDependency,
    owner: OwnerDependency,
) -> AdminUserResponse:
    """Grant a Telegram user sales-only access without creating more owners."""
    existing = await session.get(AdminUser, payload.telegram_user_id)
    if existing is not None and existing.role is AdminRole.OWNER:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT, detail="Cannot replace the owner role"
        )
    if existing is None:
        existing = AdminUser(
            telegram_user_id=payload.telegram_user_id,
            role=AdminRole.SALES,
            added_by_telegram_id=owner.telegram_user_id,
        )
        session.add(existing)
    else:
        existing.role = AdminRole.SALES
        existing.is_active = True
        existing.added_by_telegram_id = owner.telegram_user_id
    await session.commit()
    return AdminUserResponse(
        telegram_user_id=existing.telegram_user_id,
        role=existing.role.value,
        is_active=existing.is_active,
        created_at=existing.created_at,
    )


def workbook_response(workbook: Workbook, filename: str) -> Response:
    """Serialize a workbook as an XLSX download without writing personal data to disk."""
    output = BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/exports/customers")
async def export_customers(session: SessionDependency, _: OwnerDependency) -> Response:
    """Build the approved customer export on demand for the owner only."""
    rows = list(
        (
            await session.execute(
                select(Customer, LoyaltyAccount)
                .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                .order_by(Customer.created_at.desc())
            )
        ).all()
    )
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Клиенты")
    sheet.append(["ФИО", "Телефон", "Дата регистрации", "Уровень", "Баланс", "Оборот"])
    tiers = list(
        (
            await session.scalars(
                select(LoyaltyTierRule)
                .where(LoyaltyTierRule.is_active.is_(True))
                .order_by(LoyaltyTierRule.minimum_turnover)
            )
        ).all()
    )
    for customer, account in rows:
        eligible = [tier for tier in tiers if tier.minimum_turnover <= account.lifetime_turnover]
        cashback_percent = eligible[-1].cashback_percent if eligible else Decimal("0")
        sheet.append(
            [
                customer.full_name,
                customer.phone,
                customer.created_at.isoformat(),
                f"{cashback_percent}%",
                account.current_balance,
                account.lifetime_turnover,
            ]
        )
    return workbook_response(workbook, "customers.xlsx")


@router.get("/exports/purchases")
async def export_purchases(
    session: SessionDependency,
    _: OwnerDependency,
    date_from: date | None = None,
    date_to: date | None = None,
) -> Response:
    """Build the approved purchase export with selected product snapshots."""
    statement = (
        select(Purchase, Customer)
        .join(Customer, Customer.id == Purchase.customer_id)
        .options(selectinload(Purchase.items))
        .order_by(Purchase.created_at.desc())
    )
    start = as_utc_boundary(date_from, end=False)
    end = as_utc_boundary(date_to, end=True)
    if start is not None:
        statement = statement.where(Purchase.created_at >= start)
    if end is not None:
        statement = statement.where(Purchase.created_at <= end)
    rows = list((await session.execute(statement)).unique().all())

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Покупки")
    sheet.append(
        [
            "Дата",
            "Клиент",
            "Сумма",
            "Списано бонусов",
            "Начислено бонусов",
            "Кешбэк",
            "Администратор",
            "Товары",
        ]
    )
    for purchase, customer in rows:
        sheet.append(
            [
                purchase.created_at.isoformat(),
                customer.full_name,
                purchase.total_amount,
                purchase.bonus_redeemed,
                purchase.cashback_accrued,
                (
                    f"День рождения · {purchase.cashback_percent}%"
                    if purchase.cashback_source is CashbackSource.BIRTHDAY
                    else f"Уровень · {purchase.cashback_percent}%"
                ),
                purchase.recorded_by_telegram_id,
                ", ".join(item.title_snapshot for item in purchase.items),
            ]
        )
    return workbook_response(workbook, "purchases.xlsx")
