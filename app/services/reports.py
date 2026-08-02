"""In-memory XLSX reports shared by owner API and Telegram commands."""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models import CashbackSource, Customer, LoyaltyAccount, Purchase
from app.services.tier_rules import list_active_tiers


def workbook_bytes(workbook: Workbook) -> bytes:
    """Serialize an XLSX workbook without writing personal data to disk."""
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def customer_report(session: AsyncSession) -> bytes:
    """Build the canonical owner customer report used by every interface."""
    rows = list(
        (
            await session.execute(
                select(Customer, LoyaltyAccount)
                .join(LoyaltyAccount, LoyaltyAccount.customer_id == Customer.id)
                .order_by(Customer.created_at.desc())
            )
        ).all()
    )
    tiers = await list_active_tiers(session)

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Клиенты")
    sheet.append(["ФИО", "Телефон", "Дата регистрации", "Уровень", "Баланс", "Оборот"])
    for customer, account in rows:
        eligible = [tier for tier in tiers if tier.minimum_turnover <= account.lifetime_turnover]
        cashback_percent = eligible[-1].cashback_percent if eligible else Decimal("0")
        sheet.append(
            [
                customer.full_name,
                customer.phone,
                customer.created_at.isoformat(),
                f"{cashback_percent}%",
                account.current_balance,
                account.lifetime_turnover,
            ]
        )
    return workbook_bytes(workbook)


async def purchase_report(
    session: AsyncSession,
    *,
    date_from: datetime | None = None,
    date_to: datetime | None = None,
) -> bytes:
    """Build the canonical immutable-purchase report with product snapshots."""
    statement = (
        select(Purchase, Customer)
        .join(Customer, Customer.id == Purchase.customer_id)
        .options(selectinload(Purchase.items))
        .order_by(Purchase.created_at.desc())
    )
    if date_from is not None:
        statement = statement.where(Purchase.created_at >= date_from)
    if date_to is not None:
        statement = statement.where(Purchase.created_at <= date_to)
    rows = list((await session.execute(statement)).unique().all())

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet("Покупки")
    sheet.append(
        [
            "Дата",
            "Клиент",
            "Сумма",
            "Списано бонусов",
            "Начислено бонусов",
            "Кешбэк",
            "Администратор",
            "Товары",
        ]
    )
    for purchase, customer in rows:
        sheet.append(
            [
                purchase.created_at.isoformat(),
                customer.full_name,
                purchase.total_amount,
                purchase.bonus_redeemed,
                purchase.cashback_accrued,
                (
                    f"День рождения · {purchase.cashback_percent}%"
                    if purchase.cashback_source is CashbackSource.BIRTHDAY
                    else f"Уровень · {purchase.cashback_percent}%"
                ),
                purchase.recorded_by_telegram_id,
                ", ".join(item.title_snapshot for item in purchase.items),
            ]
        )
    return workbook_bytes(workbook)
